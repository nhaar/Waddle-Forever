#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# CITATION (base script reference): :contentReference[oaicite:0]{index=0}
"""
export_rooms_fast.py — Exportador optimizado (rápido y silencioso)

Cambios clave vs. tu versión original:
1) Salida de frames: SOLO 1 PNG por cada SWF, con el MISMO nombre del SWF (p.ej. `room.swf` → `room.png`).
2) Mucho más rápido:
   - Trabajo en paralelo (copiado de SWF + extracción de frame) con ThreadPoolExecutor.
   - Excel en modo write_only para escribir filas más rápido y con menor uso de RAM.
   - Menos I/O: ya no copiamos todos los PNGs temporales; solo el frame objetivo.
3) Silencioso por defecto (sin “paso por paso”):
   - Nivel de log por defecto = ERROR. Usa --log-level INFO si quieres ver progreso.
4) Compatible con tus flags originales; se añaden:
   - --workers N  → define hilos de extracción/copiar (por defecto: núm. CPUs).
   - --target-frame N → si quieres un frame distinto al 1 (por defecto 1).

Uso (idéntico + mejoras):
    python export_rooms_fast.py \
        --repo-root .. \
        --output LOST-IN-TIME/EXPORT-IMPORT \
        --engine auto \
        --ffdec "C:/Program Files/FFDec/ffdec.jar" \
        --workers 8 \
        --log-level ERROR
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
except ImportError as exc:
    raise SystemExit("openpyxl es requerido. Instálalo con: pip install openpyxl") from exc

# ---------------------------------------------------------------------------
# Import opcional del helper de extracción (si existe en ExtraerFrameSWF/)
# ---------------------------------------------------------------------------
EXTRAER_DIR = Path(__file__).resolve().parent / "ExtraerFrameSWF"
if EXTRAER_DIR.exists():
    sys.path.append(str(EXTRAER_DIR))
try:
    # Este módulo es el mismo que ya usabas; lo aprovechamos si está disponible.
    import extract_swf_frames  # type: ignore
except Exception:
    extract_swf_frames = None

# ---------------------------------------------------------------------------
# Descubrimiento de JPEXS (ffdec)
# ---------------------------------------------------------------------------
def discover_ffdec_path() -> Optional[Path]:
    def expand(candidate: Optional[str]) -> Optional[Path]:
        if not candidate:
            return None
        p = Path(candidate).expanduser()
        return p if p.exists() else None

    # Variables de entorno comunes
    for env in ("FFDEC", "FFDEC_PATH", "JPEXS_FFDEC"):
        val = expand(os.environ.get(env))
        if val:
            return val

    # Ubicaciones comunes Windows
    common = [
        Path(r"C:/Program Files/FFDec/ffdec.jar"),
        Path(r"C:/Program Files/FFDec/ffdec.exe"),
        Path(r"C:/Program Files/FFDec/ffdec.bat"),
        Path(r"C:/Program Files (x86)/FFDec/ffdec.jar"),
        Path(r"C:/Program Files (x86)/FFDec/ffdec.exe"),
        Path(r"C:/Program Files (x86)/FFDec/ffdec.bat"),
    ]
    for p in common:
        if p.exists():
            return p

    # Explorar raíces conocidas en Windows (rápido, no profundo)
    def iter_roots() -> Iterable[Path]:
        for name in ("ProgramFiles", "ProgramFiles(x86)", "ProgramW6432", "LOCALAPPDATA"):
            v = expand(os.environ.get(name))
            if v:
                yield v

    def iter_ffdec_candidates(base: Path) -> Iterable[Path]:
        names = ("ffdec.jar", "ffdec.exe", "ffdec.bat", "ffdec.cmd")
        for n in names:
            c = base / n
            if c.exists():
                yield c
        for sub in ("bin", "lib"):
            for n in names:
                c = base / sub / n
                if c.exists():
                    yield c

    for root in iter_roots():
        for name in ("JPEXS Free Flash Decompiler", "JPEXS Free Flash Decompiler (Nightly)", "FFDec", "FFDec nightly"):
            base = root / name
            if base.exists():
                for c in iter_ffdec_candidates(base):
                    return c
        for pattern in ("*/ffdec.jar", "*/ffdec.exe", "*/ffdec.bat", "*/*/ffdec.jar", "*/*/ffdec.exe", "*/*/ffdec.bat"):
            for c in root.glob(pattern):
                return c

    # Mirar si viene junto al helper
    for c in [
        EXTRAER_DIR / "ffdec.jar",
        EXTRAER_DIR / "ffdec.bat",
        EXTRAER_DIR / "ffdec.exe",
        EXTRAER_DIR / "ffdec/ffdec.jar",
        EXTRAER_DIR / "ffdec/ffdec.bat",
        EXTRAER_DIR / "ffdec/ffdec.exe",
    ]:
        if c.exists():
            return c
    return None


def normalize_ffdec_path(ffdec: Optional[Path]) -> Optional[Path]:
    if ffdec is None:
        return None
    ffdec = ffdec.expanduser()
    if ffdec.is_file():
        return ffdec
    if ffdec.is_dir():
        for n in ("ffdec.jar", "ffdec.exe", "ffdec.bat", "ffdec.cmd"):
            c = ffdec / n
            if c.exists():
                return c
        for c in ffdec.rglob("ffdec.jar"):
            return c
        for c in ffdec.rglob("ffdec.bat"):
            return c
        for c in ffdec.rglob("ffdec.exe"):
            return c
        return None
    if ffdec.suffix.lower() in {".jar", ".bat", ".cmd", ".exe"}:
        parent = ffdec.parent
        if parent.exists():
            target = ffdec.name.lower()
            for child in parent.iterdir():
                if child.name.lower() == target:
                    return child
            for n in ("ffdec.jar", "ffdec.exe", "ffdec.bat", "ffdec.cmd"):
                c = parent / n
                if c.exists():
                    return c
    return ffdec if ffdec.exists() else None

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
BETA_RELEASE = "2005-08-22"
DEFAULT_STATUS = "En proceso"
DEFAULT_CREDITS = "-"
DEFAULT_CHECKLIST = [
    {"key": "animation-test", "name": "Animation test", "checked": False, "legacy": False, "enabled": False},
    {"key": "interaction-test", "name": "Interaction test", "checked": False, "legacy": False, "enabled": False},
]
HEADERS = [
    "ID List",
    "Nombre",
    "Estado",
    "Categoría",
    "Fiesta",
    "Fechas (JSON)",
    "Sala",
    "Música",
    "Notas",
    "Ruta de imagen en Storage",
    "Checklist (JSON)",
    "Créditos",
    "Tipo de SWF",
    "Ruta de SWF en Storage",
]

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------
@dataclass
class RoomEvent:
    room: str
    start: str
    end: Optional[str]
    file_ref: Optional[str]
    category: str
    party: Optional[str] = None
    comment: Optional[str] = None
    music: Optional[int] = None
    frame: Optional[int] = None
    source: str = ""


@dataclass
class ExportRecord:
    id: int
    nombre: str
    status: str
    category: str
    fiesta: Optional[str]
    fechas_json: str
    room_label: str
    music: Optional[str]
    notes: str
    image_rel_path: str
    checklist_json: str
    credits: str
    swf_type: str
    swf_rel_path: str
    file_ref: Optional[str]
    frame: Optional[int]
    start: str

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    default_repo_root = script_dir.parent
    default_output = default_repo_root / "LOST-IN-TIME" / "EXPORT-IMPORT"

    parser = argparse.ArgumentParser(description="Exportador rápido al formato LOST-IN-TIME (1 frame por SWF, silencioso)")
    parser.add_argument("--repo-root", type=Path, default=default_repo_root, help="Ruta al repo WF")
    parser.add_argument("--output", type=Path, default=default_output, help="Carpeta de salida EXPORT-IMPORT")
    parser.add_argument("--engine", choices=["auto", "ffmpeg", "jpexs", "swfrender", "none"], default="auto",
                        help="Motor de extracción de frame. 'none' para saltar PNGs.")
    parser.add_argument("--ffdec", type=Path, help="Ruta a ffdec.jar/bat (JPEXS)")
    parser.add_argument("--preview-zoom", type=float, default=2.0, help="Zoom para JPEXS (2.0 por defecto)")
    parser.add_argument("--clean", action="store_true", help="Borra records.xlsx y storage antes de generar")
    parser.add_argument("--skip-media", action="store_true", help="No copiar SWF ni generar PNGs")
    parser.add_argument("--log-level", default="ERROR", choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                        help="Nivel de logs (por defecto ERROR para no mostrar paso a paso)")
    parser.add_argument("--workers", type=int, default=max(1, (os.cpu_count() or 4)),
                        help="Nº de hilos para copiar SWFs y extraer frames")
    parser.add_argument("--target-frame", type=int, default=1, help="Frame objetivo (1 por defecto)")
    return parser.parse_args()


def configure_logging(level: str) -> None:
    logging.basicConfig(format="[%(levelname)s] %(message)s", level=getattr(logging, level.upper(), logging.ERROR))


def slugify(value: str) -> str:
    value = value.replace("/", "-")
    value = re.sub(r"[^A-Za-z0-9_-]+", "-", value)
    value = re.sub(r"-+", "-", value)
    return value.strip("-") or "room"


def ensure_directory(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def json_dump(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False)


def parse_version(version: str) -> Tuple[date, bool]:
    parts = version.split("-")
    if len(parts) != 3:
        raise ValueError(f"Invalid version: {version}")
    year = int(parts[0])
    month_unknown = parts[1] == "##"
    day_unknown = parts[2] == "##"
    month = 1 if month_unknown else int(parts[1])
    day = 1 if day_unknown else int(parts[2])
    return date(year, month, day), (month_unknown or day_unknown)


def version_sort_key(version: str) -> Tuple[int, int, int, str]:
    d, _ = parse_version(version)
    return (d.year, d.month, d.day, version)


def compute_period_end(start: str, next_start: Optional[str]) -> Optional[str]:
    if next_start is None:
        return None
    start_date, start_unknown = parse_version(start)
    next_date, next_unknown = parse_version(next_start)
    if start_unknown or next_unknown:
        return next_start
    end_date = next_date - timedelta(days=1)
    if end_date < start_date:
        return next_start
    return end_date.isoformat()


def get_music_lookup(raw_music: Dict[str, List[Any]]) -> Dict[str, List[Tuple[str, int]]]:
    lookup: Dict[str, List[Tuple[str, int]]] = {}
    for room, entries in raw_music.items():
        if not entries:
            continue
        default_id = entries[0]
        timeline: List[Tuple[str, int]] = [("0001-01-01", default_id)]
        for entry in entries[1:]:
            timeline.append((entry["date"], entry["musicId"]))
        timeline.sort(key=lambda pair: version_sort_key(pair[0]))
        lookup[room] = timeline
    return lookup


def get_music_for_date(lookup: Dict[str, List[Tuple[str, int]]], room: str, version: str) -> Optional[int]:
    timeline = lookup.get(room)
    if not timeline:
        return None
    selected = timeline[0][1]
    version_key = version_sort_key(version)
    for entry_date, music_id in timeline:
        if version_sort_key(entry_date) <= version_key:
            selected = music_id
        else:
            break
    return selected


def call_node_for_data(repo_root: Path) -> Dict[str, Any]:
    if shutil.which("node") is None:
        raise SystemExit("Se requiere Node.js para extraer los datos del juego.")
    js_code = r"""
const { ROOMS } = require('./src/server/game-data/rooms');
const { ORIGINAL_ROOMS } = require('./src/server/game-data/release-features');
const { ROOM_OPENINGS, ROOM_UPDATES, TEMPORARY_ROOM_UPDATES, ROOM_MUSIC_TIMELINE } = require('./src/server/game-data/room-updates');
const { PARTIES } = require('./src/server/game-data/parties');
const data = {
  rooms: ROOMS,
  originalRooms: ORIGINAL_ROOMS,
  roomOpenings: ROOM_OPENINGS,
  roomUpdates: ROOM_UPDATES,
  temporaryRoomUpdates: TEMPORARY_ROOM_UPDATES,
  roomMusicTimeline: ROOM_MUSIC_TIMELINE,
  parties: PARTIES,
};
const replacer = (_key, value) => {
  if (value instanceof Map) return Object.fromEntries(value.entries());
  if (value instanceof Set) return Array.from(value.values());
  return value;
};
process.stdout.write(JSON.stringify(data, replacer));
"""
    cmd = ["node", "-r", "ts-node/register/transpile-only", "-e", js_code]
    try:
        result = subprocess.run(cmd, cwd=repo_root, check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as exc:
        logging.error("Fallo ejecutando Node: %s", exc.stderr.strip())
        raise SystemExit(1) from exc
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        logging.error("JSON inválido desde Node: %s", exc)
        raise SystemExit(1) from exc

# ---------------------------------------------------------------------------
# Ensamblado de eventos
# ---------------------------------------------------------------------------
def add_base_event(
    base_events: Dict[str, List[RoomEvent]],
    room: str,
    start: str,
    file_ref: Optional[str],
    comment: Optional[str],
    source: str,
    music_lookup: Dict[str, List[Tuple[str, int]]],
) -> None:
    if file_ref is None:
        return
    base_events[room].append(
        RoomEvent(
            room=room,
            start=start,
            end=None,
            file_ref=file_ref,
            category="base",
            comment=comment,
            source=source,
            music=get_music_for_date(music_lookup, room, start),
        )
    )


def add_temporary_event(
    temporary_events: List[RoomEvent],
    room: str,
    start: str,
    end: Optional[str],
    file_ref: Optional[str],
    category: str,
    comment: Optional[str],
    party: Optional[str],
    music: Optional[int],
    frame: Optional[int],
    source: str,
) -> None:
    if file_ref is None:
        return
    if end is not None and start == end:
        return
    temporary_events.append(
        RoomEvent(
            room=room,
            start=start,
            end=end,
            file_ref=file_ref,
            category=category,
            party=party,
            comment=comment,
            music=music,
            frame=frame,
            source=source,
        )
    )


def get_sub_update_dates(update: Dict[str, Any], index: int) -> Tuple[str, Optional[str]]:
    updates = update.get("updates")
    if not updates:
        raise ValueError("Update sin sub-updates")
    sub = updates[index]
    end = sub.get("end")
    if end is None:
        next_update = updates[index + 1] if index + 1 < len(updates) else None
        if next_update and next_update.get("date"):
            end = next_update["date"]
    if end is None:
        end = update.get("end")
    start = sub.get("date", update.get("date"))
    if start == end:
        return start, None
    return start, end


def process_parties(
    parties: List[Dict[str, Any]],
    base_events: Dict[str, List[RoomEvent]],
    temporary_events: List[RoomEvent],
    music_lookup: Dict[str, List[Tuple[str, int]]],
) -> None:
    for party in parties:
        party_name = party.get("name", "") or None
        party_start = party["date"]
        party_end = party.get("end")
        party_category = "event" if party.get("event") else "party"
        room_frames = party.get("roomFrames", {})
        room_music = party.get("music", {})
        party_comment = party.get("roomComment")

        permanent = party.get("permanentChanges") or {}
        for room, file_ref in (permanent.get("roomChanges") or {}).items():
            add_base_event(
                base_events, room, party_start, file_ref, permanent.get("roomComment") or party_comment,
                f"party-permanent:{party_name or ''}", music_lookup
            )

        for room, file_ref in (party.get("roomChanges") or {}).items():
            add_temporary_event(
                temporary_events, room, party_start, party_end, file_ref, party_category, party_comment,
                party_name, room_music.get(room), room_frames.get(room), "party"
            )

        updates = party.get("updates") or []
        for idx, update in enumerate(updates):
            sub_start, sub_end = get_sub_update_dates(party, idx)
            update_frames = update.get("roomFrames", {})
            update_music = update.get("music", {})
            for room, file_ref in (update.get("roomChanges") or {}).items():
                add_temporary_event(
                    temporary_events, room, sub_start, sub_end, file_ref, party_category,
                    update.get("comment") or party_comment, party_name,
                    update_music.get(room), update_frames.get(room), "party-update"
                )

        construction = party.get("construction")
        if construction:
            const_start = construction["date"]
            const_end = party_start
            const_frames = construction.get("roomFrames", {})
            const_music = construction.get("music", {})
            for room, file_ref in (construction.get("changes") or {}).items():
                add_temporary_event(
                    temporary_events, room, const_start, const_end, file_ref, "construction",
                    construction.get("comment"), party_name, const_music.get(room), const_frames.get(room), "construction"
                )
            const_updates = construction.get("updates") or []
            fake_parent = {"date": const_start, "end": const_end, "updates": const_updates}
            for idx, update in enumerate(const_updates):
                sub_start, sub_end = get_sub_update_dates(fake_parent, idx)
                for room, file_ref in (update.get("changes") or {}).items():
                    add_temporary_event(
                        temporary_events, room, sub_start, sub_end, file_ref, "construction",
                        update.get("comment"), party_name, None, None, "construction-update"
                    )

        consequences = party.get("consequences") or {}
        for room, file_ref in (consequences.get("roomChanges") or {}).items():
            add_base_event(
                base_events, room, party_end or party_start, file_ref,
                consequences.get("roomComment") or party_comment,
                f"party-consequence:{party_name or ''}", music_lookup
            )


def build_events(game_data: Dict[str, Any]) -> Tuple[List[RoomEvent], Dict[str, List[RoomEvent]]]:
    music_lookup = get_music_lookup(game_data["roomMusicTimeline"])
    base_events: Dict[str, List[RoomEvent]] = defaultdict(list)
    temporary_events: List[RoomEvent] = []

    for room, file_ref in game_data["originalRooms"].items():
        add_base_event(base_events, room, BETA_RELEASE, file_ref, None, "original", music_lookup)

    for opening in game_data["roomOpenings"]:
        comment = opening.get("comment")
        if opening.get("fileRef"):
            add_base_event(base_events, opening["room"], opening["date"], opening["fileRef"], comment, "opening", music_lookup)
        for room, file_ref in (opening.get("otherRooms") or {}).items():
            add_base_event(base_events, room, opening["date"], file_ref, comment, "opening-other", music_lookup)

    for room, updates in game_data["roomUpdates"].items():
        for update in updates:
            add_base_event(base_events, room, update["date"], update["fileRef"], update.get("comment"), "update", music_lookup)

    for room, updates in (game_data["temporaryRoomUpdates"] or {}).items():
        for update in updates:
            add_temporary_event(
                temporary_events, room, update["date"], update.get("end"), update["fileRef"],
                "temporary", update.get("comment"), None, None, update.get("frame"), "temporary"
            )

    process_parties(game_data["parties"], base_events, temporary_events, music_lookup)

    for room, events in base_events.items():
        events.sort(key=lambda ev: version_sort_key(ev.start))
        for idx, event in enumerate(events):
            next_start = events[idx + 1].start if idx + 1 < len(events) else None
            event.end = compute_period_end(event.start, next_start)

    return temporary_events, base_events


def build_records(rooms: Dict[str, Any], temporary_events: List[RoomEvent], base_events: Dict[str, List[RoomEvent]]) -> List[ExportRecord]:
    all_events: List[RoomEvent] = []
    for events in base_events.values():
        all_events.extend(events)
    all_events.extend(temporary_events)
    all_events.sort(key=lambda ev: (version_sort_key(ev.start), ev.room, ev.category))

    records: List[ExportRecord] = []
    for idx, event in enumerate(all_events, start=1):
        room_info = rooms.get(event.room) or {}
        room_label = room_info.get("name", event.room.title())
        file_ref = event.file_ref
        if file_ref:
            prefix, path_part = file_ref.split(":", 1)
            file_name = Path(path_part).name                         # p.ej. room.swf
            swf_stem = Path(path_part).stem                          # p.ej. room
            nombre = slugify(Path(path_part).stem)
            swf_type = prefix
            image_rel_path = f"record-images/{idx}/{swf_stem}.png"   # MISMO nombre que el SWF
            swf_rel_path = f"swf-files/{idx}/{file_name}"
        else:
            nombre = slugify(f"{event.room}-{event.start}")
            swf_type = ""
            swf_rel_path = ""
            image_rel_path = f"record-images/{idx}/{nombre}.png"     # Fallback si no hay SWF

        fechas_json = json_dump([{"start": event.start, "end": event.end}])
        checklist_json = json_dump([{**item, "key": f"{nombre}-{item['key']}"} for item in DEFAULT_CHECKLIST])
        music_value = event.music if event.music is not None else None
        notes = event.comment or (f"fileRef: '{file_ref}'" if file_ref else "")

        records.append(
            ExportRecord(
                id=idx,
                nombre=nombre,
                status=DEFAULT_STATUS,
                category=event.category,
                fiesta=event.party,
                fechas_json=fechas_json,
                room_label=room_label,
                music=str(music_value) if music_value is not None else None,
                notes=notes,
                image_rel_path=image_rel_path,
                checklist_json=checklist_json,
                credits=DEFAULT_CREDITS,
                swf_type=swf_type,
                swf_rel_path=swf_rel_path,
                file_ref=file_ref,
                frame=event.frame,
                start=event.start,
            )
        )
    return records

# ---------------------------------------------------------------------------
# Salida: Excel + media
# ---------------------------------------------------------------------------
def write_excel(output_path: Path, records: List[ExportRecord]) -> None:
    # Modo write_only: mucho más rápido y consume menos memoria
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Registros", 0)
    ws.append(HEADERS)
    for r in records:
        ws.append([
            str(r.id), r.nombre, r.status, r.category, r.fiesta, r.fechas_json, r.room_label, r.music, r.notes,
            r.image_rel_path, r.checklist_json, r.credits, r.swf_type, r.swf_rel_path
        ])
    # openpyxl en write_only crea una hoja vacía por defecto; elimínala si aparece
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    wb.save(output_path)


def resolve_media_source(repo_root: Path, file_ref: str) -> Optional[Path]:
    if ":" not in file_ref:
        return None
    prefix, relative = file_ref.split(":", 1)
    return repo_root / "media" / "default" / prefix / relative


def extract_single_frame(
    swf_path: Path,
    dest_dir: Path,
    out_png_name: str,   # <<--- MISMO nombre que el SWF, pero .png
    engine: str,
    target_frame: int,
    ffdec: Optional[Path],
    preview_zoom: Optional[float],
) -> None:
    """
    Extrae SOLO UN PNG del SWF:
    - Prioriza sacar el frame `target_frame` (default 1).
    - No deja archivos extra; solo {dest_dir}/{out_png_name}.
    - Evita “paso por paso” (usa logging y subprocess en silencio).
    """
    if engine == "none":
        return
    if not swf_path.exists():
        return

    # Descubrir ffdec si hace falta
    def available_ffdec() -> Optional[Path]:
        if ffdec and ffdec.exists():
            return ffdec
        auto = discover_ffdec_path()
        return auto if (auto and auto.exists()) else None

    # Explorar frames detectados en un tmpdir
    def discover_first_frame(root: Path) -> Optional[Path]:
        # Preferimos frame_00001.png si existe; si no, el primer .png
        candidates = sorted(root.rglob("frame_*.png"))
        if candidates:
            return next((c for c in candidates if c.name.lower() == f"frame_{target_frame:05d}.png"), candidates[0])
        any_pngs = sorted(root.rglob("*.png"))
        return any_pngs[0] if any_pngs else None

    def try_ffmpeg(tmp: Path) -> Optional[Path]:
        if shutil.which("ffmpeg") is None:
            return None
        # ffmpeg puede sacar 1 frame con -vframes 1 (rápido cuando SWF es soportado)
        out_path = tmp / "ffmpeg_frame.png"
        cmd = ["ffmpeg", "-y", "-loglevel", "error", "-i", str(swf_path), "-frames:v", "1", str(out_path)]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return out_path if out_path.exists() else None
        except Exception:
            return None

    def try_swfrender(tmp: Path) -> Optional[Path]:
        # swfrender + swfdump (SWFTools) — no siempre instalado
        if not (shutil.which("swfrender") and shutil.which("swfdump")):
            return None
        out_path = tmp / "swfrender_frame.png"
        # Algunos swfrender no soportan pedir frame específico; sacamos el primero
        cmd = ["swfrender", str(swf_path), "-o", str(out_path)]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return out_path if out_path.exists() else None
        except Exception:
            return None

    def try_jpexs(tmp: Path) -> Optional[Path]:
        # Si tenemos helper, lo usamos (suele exportar todos los frames).
        # Tras exportar, elegimos el frame objetivo y borramos el resto.
        exported_frame = None
        if extract_swf_frames is not None:
            # exportación genérica (puede volcar varios PNGs)
            try:
                extract_swf_frames.jpexs_extract(  # type: ignore[attr-defined]
                    str(swf_path), str(tmp), str(available_ffdec() or ""), zoom=preview_zoom
                )
                exported_frame = discover_first_frame(tmp)
            except Exception:
                exported_frame = None
        else:
            # CLI directa de ffdec (si existe). Intentamos rango 1-1 si es soportado.
            ffdec_exec = available_ffdec()
            if ffdec_exec:
                out_dir = tmp / "ffdec"
                ensure_directory(out_dir)
                # Intento con opciones comunes de CLI:
                # -cli -format png -export frame -range 1-1 <outdir> <file.swf>
                # (algunas builds usan otra sintaxis; si falla, ignoramos)
                cmd = [str(ffdec_exec), "-cli", "-format", "png", "-export", "frame", "-range", f"{target_frame}-{target_frame}", str(out_dir), str(swf_path)]
                try:
                    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    exported_frame = discover_first_frame(out_dir)
                except Exception:
                    # plan B: sin rango, exporta todo y elegimos:
                    cmd2 = [str(ffdec_exec), "-cli", "-format", "png", "-export", "frame", str(out_dir), str(swf_path)]
                    try:
                        subprocess.run(cmd2, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                        exported_frame = discover_first_frame(out_dir)
                    except Exception:
                        exported_frame = None
        return exported_frame

    with TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        chosen: Optional[Path] = None

        engines = []
        if engine == "auto":
            engines = ["ffmpeg", "jpexs", "swfrender"]
        else:
            engines = [engine]

        for eng in engines:
            if eng == "ffmpeg":
                chosen = try_ffmpeg(tmp)
            elif eng == "jpexs":
                chosen = try_jpexs(tmp)
            elif eng == "swfrender":
                chosen = try_swfrender(tmp)
            else:
                chosen = None
            if chosen and chosen.exists():
                break

        if not chosen:
            return

        ensure_directory(dest_dir)
        final_png = dest_dir / out_png_name
        if final_png.exists():
            try:
                final_png.unlink()
            except Exception:
                pass
        shutil.copy2(chosen, final_png)


def copy_assets_parallel(
    repo_root: Path,
    output_root: Path,
    records: List[ExportRecord],
    engine: str,
    ffdec: Optional[Path],
    preview_zoom: Optional[float],
    workers: int,
    target_frame: int,
) -> None:
    storage_root = output_root / "storage"
    swf_root = storage_root / "swf-files"
    img_root = storage_root / "record-images"
    ffdec_exec = ffdec if ffdec and ffdec.exists() else None
    if ffdec and not ffdec_exec:
        logging.warning("ffdec configurado pero no accesible: %s", ffdec)

    def process_record(rec: ExportRecord) -> None:
        if not rec.file_ref:
            return
        source = resolve_media_source(repo_root, rec.file_ref)
        if source is None or not source.exists():
            return
        # Copiar SWF
        dest_swf_dir = swf_root / str(rec.id)
        ensure_directory(dest_swf_dir)
        dest_swf = dest_swf_dir / Path(rec.swf_rel_path).name
        if not dest_swf.exists():
            shutil.copy2(source, dest_swf)
        # Generar SOLO 1 PNG con el mismo nombre del SWF (stem + .png)
        dest_img_dir = img_root / str(rec.id)
        out_png_name = Path(rec.swf_rel_path).stem + ".png"
        extract_single_frame(dest_swf, dest_img_dir, out_png_name, engine, target_frame, ffdec_exec, preview_zoom)

    tasks = [rec for rec in records if rec.file_ref]
    if workers <= 1:
        for rec in tasks:
            process_record(rec)
    else:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = [ex.submit(process_record, rec) for rec in tasks]
            for _ in as_completed(futures):
                pass  # silencioso: no mostramos progreso

def clean_output(output_root: Path) -> None:
    records_path = output_root / "records.xlsx"
    storage_path = output_root / "storage"
    if records_path.exists():
        records_path.unlink()
    if storage_path.exists():
        shutil.rmtree(storage_path)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    args = parse_args()
    configure_logging(args.log_level)

    if args.ffdec is None:
        auto_ffdec = discover_ffdec_path()
        if auto_ffdec:
            args.ffdec = auto_ffdec
    if args.ffdec is not None:
        args.ffdec = normalize_ffdec_path(args.ffdec)

    repo_root = args.repo_root.resolve()
    output_root = args.output.resolve()
    ensure_directory(output_root)

    if args.clean:
        clean_output(output_root)

    # 1) Datos del juego
    game_data = call_node_for_data(repo_root)

    # 2) Eventos → Registros
    temporary_events, base_events = build_events(game_data)
    records = build_records(game_data["rooms"], temporary_events, base_events)

    # 3) Excel rápido
    records_path = output_root / "records.xlsx"
    write_excel(records_path, records)

    # 4) Media (paralelo, 1 frame por SWF con mismo nombre)
    if not args.skip_media:
        copy_assets_parallel(
            repo_root=repo_root,
            output_root=output_root,
            records=records,
            engine=args.engine,
            ffdec=args.ffdec,
            preview_zoom=args.preview_zoom,
            workers=max(1, int(args.workers)),
            target_frame=max(1, int(args.target_frame)),
        )

if __name__ == "__main__":
    main()
